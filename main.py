import asyncio
import os
import sys
from copilot import CopilotClient
from copilot.tools import define_tool
from copilot.generated.session_events import SessionEventType
from pydantic import BaseModel, Field

MODEL = "claude-opus-4.6-fast"
TIME_OUT = 3000

# 出力先ディレクトリ（実行時に設定）
OUTPUT_DIR = "."


# ══════════════════════════════════════════════════════════════════
# Tool Definitions
# ══════════════════════════════════════════════════════════════════

class SaveMarkdownParams(BaseModel):
    filename: str = Field(description="Output filename")
    content: str = Field(description="Markdown content")


@define_tool(description="マークダウンコンテンツをファイルに保存します。")
async def save_markdown_file(params: SaveMarkdownParams) -> dict:
    filepath = os.path.join(OUTPUT_DIR, params.filename)
    dirpath = os.path.dirname(filepath)
    if dirpath:
        os.makedirs(dirpath, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(params.content)
    return {"saved": filepath, "bytes": os.path.getsize(filepath)}


# ══════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════

def create_streaming_handler():
    chunks, done = [], asyncio.Event()

    def handler(event):
        et = event.type
        d = event.data

        if et == SessionEventType.ASSISTANT_MESSAGE_DELTA:
            t = d.delta_content or ""
            sys.stdout.write(t)
            sys.stdout.flush()
            chunks.append(t)

        elif et == SessionEventType.TOOL_EXECUTION_START:
            name = d.tool_name or "unknown"
            args = d.arguments or ""
            print(f"\n  🔧 ツール呼び出し: {name}({_truncate(str(args), 100)})")

        elif et == SessionEventType.TOOL_EXECUTION_PROGRESS:
            msg = d.progress_message or ""
            if msg:
                print(f"  ⏳ {msg}")

        elif et == SessionEventType.TOOL_EXECUTION_COMPLETE:
            name = d.tool_name or "unknown"
            print(f"  ✅ {name} 完了")

        elif et == SessionEventType.ASSISTANT_TURN_START:
            print("  🤖 Copilot 思考中...")

        elif et == SessionEventType.ASSISTANT_TURN_END:
            pass  # ターン終了（静かに処理）

        elif et == SessionEventType.SESSION_ERROR:
            msg = d.message or d.error_type or "不明なエラー"
            print(f"\n  ❌ エラー: {msg}")

        elif et == SessionEventType.SESSION_IDLE:
            done.set()

    return handler, chunks, done


def _truncate(s: str, maxlen: int) -> str:
    return s if len(s) <= maxlen else s[:maxlen] + "..."


SYSTEM_MESSAGE = (
    "あなたはドキュメント変換の専門家です。指定されたファイルの内容を高品質なマークダウンに変換します。\n\n"
    "## ルール\n"
    "- テキストは適切なマークダウン書式（見出し、太字、リスト等）で出力\n"
    "- 画像が含まれる場合、画像の**詳細な説明**をマークダウンで記述する（画像ファイル自体は不要。mermaid図での説明も可）\n"
    "- 図形オブジェクトによる図がある場合、オブジェクト構造・位置関係を分析し **mermaid形式の図** で再現する\n"
    "  - 複雑な図の場合は **複数のmermaid図** に分割してもよい\n"
    "- テーブルはマークダウンテーブルで出力\n"
    "- 出力は日本語\n"
    "- 必ず save_markdown_file ツールを使って結果をファイルに保存すること\n"
)

ALL_TOOLS = [save_markdown_file]


# ══════════════════════════════════════════════════════════════════
# Processors
# ══════════════════════════════════════════════════════════════════

async def process_pptx(client, path, basename):
    print(f"\n📊 PPTX を処理します...\n")

    session = await client.create_session({
        "model": MODEL,
        "streaming": True,
        "tools": ALL_TOOLS,
        "system_message": {"content": SYSTEM_MESSAGE},
    })

    h, _, done = create_streaming_handler()
    session.on(h)

    prompt = (
        f"以下のPPTXファイルを読み取り、スライドごとに個別のマークダウンファイルとして保存してください。\n"
        f"marmaidは文法に間違いが無いかチェックを実施してから保存するようにしてください。\n\n"
        f"ファイルパス（絶対パス）: {path}\n\n"
        f"## 出力ルール\n"
        f"- 1スライドにつき1つのマークダウンファイルを作成\n"
        f"- ファイル名の形式: {basename}_{{スライド番号}}_mdize.md\n"
        f"  例: {basename}_1_mdize.md, {basename}_2_mdize.md, ...\n\n"
        f"## 変換ルール\n"
        f"- テキストは見出し・太字・リスト等で適切にフォーマット\n"
        f"- 画像があれば詳細に説明（mermaid図での補足も可）\n"
        f"- 図形オブジェクトがある場合、位置関係を分析しmermaid図で再現\n"
        f"- テーブルはマークダウンテーブルで出力\n"
        f"- 各スライドを save_markdown_file ツールで個別に保存\n"
    )

    await session.send_and_wait({"prompt": prompt}, timeout=TIME_OUT)
    await done.wait()
    print()


async def process_pptx_split(client, path, basename):
    from pptx import Presentation

    n = len(Presentation(path).slides)
    print(f"\n📊 PPTX（分割分析）: 全{n}スライドを個別に処理します...\n")

    for i in range(1, n + 1):
        print(f"{'─' * 50}")
        print(f"🔄 スライド {i}/{n}")
        print(f"{'─' * 50}")

        fname = f"{basename}_{i}_mdize.md"
        session = await client.create_session({
            "model": MODEL,
            "streaming": True,
            "tools": ALL_TOOLS,
            "system_message": {"content": SYSTEM_MESSAGE},
        })

        h, _, done = create_streaming_handler()
        session.on(h)

        prompt = (
            f"以下のPPTXファイルのスライド{i}（全{n}スライド中）を読み取り、マークダウンに変換してください。\n"
            f"このスライドだけを詳細に分析してください。\n"
            f"mermaidは文法に間違いが無いかチェックを実施してから保存するようにしてください。\n\n"
            f"ファイルパス（絶対パス）: {path}\n"
            f"対象スライド番号: {i}\n"
            f"出力ファイル名: {fname}\n\n"
            f"## 変換ルール\n"
            f"- テキストは見出し・太字・リスト等で適切にフォーマット\n"
            f"- 画像があれば詳細に説明（mermaid図での補足も可）\n"
            f"- 図形オブジェクトがある場合、位置関係を分析しmermaid図で再現\n"
            f"- テーブルはマークダウンテーブルで出力\n"
            f"- save_markdown_file ツールで '{fname}' に保存\n"
        )

        await session.send_and_wait({"prompt": prompt}, timeout=TIME_OUT)
        await done.wait()
        print(f"\n💾 保存完了: {os.path.join(OUTPUT_DIR, fname)}\n")


async def process_docx(client, path, basename):
    print(f"\n📝 DOCX を処理します...\n")

    fname = f"{basename}_mdize.md"

    session = await client.create_session({
        "model": MODEL,
        "streaming": True,
        "tools": ALL_TOOLS,
        "system_message": {"content": SYSTEM_MESSAGE},
    })

    h, _, done = create_streaming_handler()
    session.on(h)

    prompt = (
        f"以下のDOCXファイルを読み取り、マークダウンに変換してください。\n"
        f"marmaidは文法に間違いが無いかチェックを実施してから保存するようにしてください。\n\n"
        f"ファイルパス（絶対パス）: {path}\n"
        f"出力ファイル名: {fname}\n\n"
        f"## 変換ルール\n"
        f"- 見出しスタイル (Heading 1, 2 等) は # 見出しに変換\n"
        f"- 画像があれば詳細に説明（mermaid図での補足も可）\n"
        f"- 図形オブジェクトがある場合、mermaid図で再現\n"
        f"- テーブルはマークダウンテーブルで出力\n"
        f"- save_markdown_file ツールで '{fname}' に保存\n"
    )

    await session.send_and_wait({"prompt": prompt}, timeout=TIME_OUT)
    await done.wait()
    print()


async def process_docx_split(client, path, basename):
    from docx import Document

    doc = Document(path)
    # セクション（見出し1）単位で分割対象を特定
    sections = []
    for i, para in enumerate(doc.paragraphs):
        if para.style and para.style.name.startswith("Heading 1"):
            sections.append({"index": len(sections) + 1, "title": para.text.strip() or f"Section{len(sections) + 1}"})

    if not sections:
        # 見出し1が無い場合は通常処理にフォールバック
        print("  ℹ️  Heading 1 が見つからないため、通常モードで処理します。")
        await process_docx(client, path, basename)
        return

    n = len(sections)
    print(f"\n📝 DOCX（分割分析）: {n}セクションを個別に処理します...\n")

    for sec in sections:
        idx = sec["index"]
        title = sec["title"]
        safe_title = title.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_")[:30]
        fname = f"{basename}_{idx}_{safe_title}_mdize.md"

        print(f"{'─' * 50}")
        print(f"🔄 セクション {idx}/{n}: '{title}'")
        print(f"{'─' * 50}")

        session = await client.create_session({
            "model": MODEL,
            "streaming": True,
            "tools": ALL_TOOLS,
            "system_message": {"content": SYSTEM_MESSAGE},
        })

        h, _, done = create_streaming_handler()
        session.on(h)

        prompt = (
            f"以下のDOCXファイルのセクション「{title}」（Heading 1 の第{idx}セクション、全{n}セクション中）を読み取り、\n"
            f"そのセクションの内容だけをマークダウンに変換してください。\n"
            f"mermaidは文法に間違いが無いかチェックを実施してから保存するようにしてください。\n\n"
            f"ファイルパス（絶対パス）: {path}\n"
            f"出力ファイル名: {fname}\n\n"
            f"## 変換ルール\n"
            f"- 見出しスタイル (Heading 1, 2 等) は # 見出しに変換\n"
            f"- 画像があれば詳細に説明（mermaid図での補足も可）\n"
            f"- 図形オブジェクトがある場合、mermaid図で再現\n"
            f"- テーブルはマークダウンテーブルで出力\n"
            f"- save_markdown_file ツールで '{fname}' に保存\n"
        )

        await session.send_and_wait({"prompt": prompt}, timeout=TIME_OUT)
        await done.wait()
        print(f"\n💾 保存完了: {os.path.join(OUTPUT_DIR, fname)}\n")


async def process_xlsx(client, path, basename):
    print(f"\n📊 XLSX を処理します...\n")

    session = await client.create_session({
        "model": MODEL,
        "streaming": True,
        "tools": ALL_TOOLS,
        "system_message": {"content": SYSTEM_MESSAGE},
    })

    h, _, done = create_streaming_handler()
    session.on(h)

    prompt = (
        f"以下のXLSXファイルを読み取り、シートごとに個別のマークダウンファイルとして保存してください。\n"
        f"marmaidは文法に間違いが無いかチェックを実施してから保存するようにしてください。\n\n"
        f"ファイルパス（絶対パス）: {path}\n\n"
        f"## 出力ルール\n"
        f"- 1シートにつき1つのマークダウンファイルを作成\n"
        f"- ファイル名の形式: {basename}_{{シート番号}}_{{シート名}}_mdize.md\n"
        f"  例: {basename}_1_Sheet1_mdize.md\n\n"
        f"## 変換ルール\n"
        f"- データはマークダウンテーブルで出力\n"
        f"- チャートがあれば説明を記述\n"
        f"- 画像があれば詳細に説明\n"
        f"- 各シートを save_markdown_file ツールで個別に保存\n"
    )

    await session.send_and_wait({"prompt": prompt}, timeout=TIME_OUT)
    await done.wait()
    print()


async def process_xlsx_split(client, path, basename):
    from openpyxl import load_workbook

    sheets = load_workbook(path, data_only=True).sheetnames
    n = len(sheets)
    print(f"\n📊 XLSX（分割分析）: 全{n}シートを個別に処理します...\n")

    for idx, name in enumerate(sheets, 1):
        print(f"{'─' * 50}")
        print(f"🔄 シート {idx}/{n}: '{name}'")
        print(f"{'─' * 50}")

        safe = name.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_")
        fname = f"{basename}_{idx}_{safe}_mdize.md"

        session = await client.create_session({
            "model": MODEL,
            "streaming": True,
            "tools": ALL_TOOLS,
            "system_message": {"content": SYSTEM_MESSAGE},
        })

        h, _, done = create_streaming_handler()
        session.on(h)

        prompt = (
            f"以下のXLSXファイルのシート '{name}'（全{n}シート中の第{idx}シート）を読み取り、\n"
            f"このシートだけを詳細に分析してマークダウンに変換してください。\n"
            f"mermaidは文法に間違いが無いかチェックを実施してから保存するようにしてください。\n\n"
            f"ファイルパス（絶対パス）: {path}\n"
            f"対象シート名: {name}\n"
            f"出力ファイル名: {fname}\n\n"
            f"## 変換ルール\n"
            f"- データはマークダウンテーブルで出力\n"
            f"- チャートがあれば説明を記述\n"
            f"- 画像があれば詳細に説明\n"
            f"- save_markdown_file ツールで '{fname}' に保存\n"
        )

        await session.send_and_wait({"prompt": prompt}, timeout=TIME_OUT)
        await done.wait()
        print(f"\n💾 保存完了: {os.path.join(OUTPUT_DIR, fname)}\n")


# ══════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════

async def main():
    global OUTPUT_DIR

    print("=" * 60)
    print("📄 mdize — ファイルをマークダウンに変換")
    print("=" * 60)
    print("\n変換するファイルのパスを入力してください (pptx / docx / xlsx):")
    file_path = input("> ").strip().strip('"').strip("'")

    if not os.path.isfile(file_path):
        print(f"❌ ファイルが見つかりません: {file_path}")
        return

    file_path = os.path.abspath(file_path)
    ext = os.path.splitext(file_path)[1].lower()
    basename = os.path.splitext(os.path.basename(file_path))[0]

    if ext not in (".pptx", ".docx", ".xlsx"):
        print(f"❌ 未対応のファイル形式: {ext}  (対応: .pptx .docx .xlsx)")
        return

    OUTPUT_DIR = os.path.dirname(file_path)
    print(f"\n✅ ファイル: {file_path}")
    print(f"📁 出力先:  {OUTPUT_DIR}\n")

    # 分割分析モードの確認
    print("分割分析しますか？（スライド/ページ/シート単位で個別にCopilotに依頼します）")
    split_choice = input("(y/N): ").strip().lower()
    split_mode = split_choice in ("y", "yes")
    if split_mode:
        print("  → 分割分析モードで実行します\n")
    else:
        print("  → 一括処理モードで実行します\n")

    client = CopilotClient()
    await client.start()

    try:
        if ext == ".pptx":
            if split_mode:
                await process_pptx_split(client, file_path, basename)
            else:
                await process_pptx(client, file_path, basename)
        elif ext == ".docx":
            if split_mode:
                await process_docx_split(client, file_path, basename)
            else:
                await process_docx(client, file_path, basename)
        elif ext == ".xlsx":
            if split_mode:
                await process_xlsx_split(client, file_path, basename)
            else:
                await process_xlsx(client, file_path, basename)
    finally:
        await client.stop()

    print("=" * 60)
    print("✅ 変換完了!")
    print(f"📁 出力先: {OUTPUT_DIR}")
    print("=" * 60)


asyncio.run(main())